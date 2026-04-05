#!/usr/bin/env python3
"""Build a polished reproducibility bundle for the OSD-679 age-prediction paper.

Outputs a small GitHub-friendly package with:
- Excel workbooks
- CSV companions
- README and manifest JSON

The goal is to expose image-to-age mapping, primary benchmark subsets, split
definitions, and key results in a compact, manuscript-oriented form.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "reproducibility" / "osd679_age_prediction_release"
CSV_DIR = OUT_DIR / "csv"

IMAGE_AGE_MAP = ROOT / "metadata" / "image_age_mapping.csv"
S1 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S1_Cohort_Characteristics_and_Sample_Counts.csv"
S2 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S2_Control_Model_Performance_3Fold_CV.csv"
S3 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S3_Control_Test_Performance_By_Cohort_Day.csv"
S4 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S4_Backbone_Ablation_Fair_Protocol.csv"
S5 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S5_LoRA_Ablation_Full_Results.csv"
S6 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S6_Feature_Distillation_Ablation_Summary.csv"
S7 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S7_InterEye_Reliability_Control_Animals.csv"
S8 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S8_Hyperparameters_Optuna_Best_Trial.csv"
S9 = ROOT / "outputs" / "paper1" / "supplementary" / "tables" / "Supplementary_Table_S9_CrossValidation_Split_Definitions_Rat_IDs.csv"
TABLE3_MAIN = ROOT / "outputs" / "paper1" / "tables" / "table3_backbone_ablation_mainpaper.csv"

BEST_WORST_MANIFEST = ROOT / "outputs" / "paper1" / "control_best_worst_magma_xception" / "selected_sample_manifest.csv"
BEST_WORST_IMAGES = ROOT / "outputs" / "paper1" / "control_best_worst_magma_xception" / "selected_images_rows.csv"
BEST_WORST_REVIEW = ROOT / "outputs" / "paper1" / "control_best_worst_magma_xception" / "review_by_sample_index.csv"

WORKBOOK_1 = OUT_DIR / "Supplementary_Data_1_Image_to_Age_Mapping.xlsx"
WORKBOOK_2 = OUT_DIR / "Supplementary_Data_2_Benchmark_Splits_and_Results.xlsx"
WORKBOOK_3 = OUT_DIR / "Supplementary_Data_3_Qualitative_Examples.xlsx"
MANIFEST_JSON = OUT_DIR / "bundle_manifest.json"
README_MD = OUT_DIR / "README.md"


HEADER_FILL = PatternFill("solid", fgColor="274C77")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")


def relpath_or_keep(path_str: str) -> str:
    if not isinstance(path_str, str) or path_str == "":
        return path_str
    p = Path(path_str)
    try:
        return str(p.relative_to(ROOT))
    except Exception:
        return path_str


def normalize_group(series: pd.Series, baseline_flags: pd.Series | None = None) -> pd.Series:
    out = series.fillna("").astype(str).str.strip().replace({"HLS_U": "HLS (U)"})
    if baseline_flags is not None:
        baseline_mask = out.eq("") & baseline_flags.fillna(False)
        out.loc[baseline_mask] = "Baseline"
    out = out.replace({"": "Unknown"})
    return out


def clean_image_age_mapping() -> pd.DataFrame:
    df = pd.read_csv(IMAGE_AGE_MAP)
    df["cohort"] = df["cohort"].astype(str).str.strip()
    df = df[df["cohort"].isin(["1", "2", "3"])].copy()
    df["group"] = normalize_group(df["group_from_path"], df.get("is_baseline_path"))
    df["image_path_relative"] = df["image_path"].map(relpath_or_keep)
    rename_map = {
        "final_age_days": "chronological_age_days",
        "base_age_days": "base_age_days",
        "day_component_days": "day_offset_days",
        "cohort_name": "cohort_label",
        "image_type": "oct_image_type",
        "material_type": "eye_label",
        "subfolder_kind": "session_label",
        "sample_name": "sample_id",
        "folder_day_raw": "folder_day",
        "bioptigen_subfolder_raw": "bioptigen_subfolder",
    }
    keep = [
        "image_path_relative",
        "rat_id",
        "sample_name",
        "eye",
        "cohort",
        "cohort_name",
        "group",
        "day",
        "final_age_days",
        "base_age_days",
        "day_component_days",
        "sex",
        "image_type",
        "material_type",
        "subfolder_kind",
        "folder_day_raw",
        "bioptigen_subfolder_raw",
        "class_label",
    ]
    out = df[keep].rename(columns=rename_map)
    out["benchmark_age_days"] = pd.to_numeric(out["base_age_days"], errors="coerce") + pd.to_numeric(out["day"], errors="coerce")
    out = out.sort_values(["cohort", "group", "rat_id", "day", "eye", "oct_image_type", "image_path_relative"]).reset_index(drop=True)
    return out


def build_primary_subsets(df_all: pd.DataFrame, fold_map: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    fold_lookup = fold_map[["rat_id", "fold"]].drop_duplicates().rename(columns={"fold": "control_cv_fold"})
    oct_ok = df_all["oct_image_type"].isin(["BScanThumb", "REGAVG"])

    eligible = df_all[oct_ok].copy()
    controls = eligible[(eligible["group"] == "Controls") & eligible["day"].isin([0, 90])].copy()
    controls = controls.merge(fold_lookup, on="rat_id", how="left")
    allgroups = eligible[(eligible["group"].isin(["Controls", "HLS (U)"])) & eligible["day"].isin([0, 90])].copy()
    allgroups = allgroups.merge(fold_lookup, on="rat_id", how="left")

    summary_rows = []
    for name, sub in {
        "all_mapped_images_c123": df_all,
        "oct_agepred_eligible_c123": eligible,
        "paper1_controls_day0_day90": controls,
        "paper1_controls_hls_day0_day90": allgroups,
    }.items():
        summary_rows.append(
            {
                "subset_name": name,
                "n_rows": int(len(sub)),
                "n_unique_rats": int(sub["rat_id"].nunique()),
                "n_unique_eyes": int(sub[["rat_id", "eye"]].drop_duplicates().shape[0]),
                "n_unique_rat_eye_day": int(sub[["rat_id", "eye", "day"]].drop_duplicates().shape[0]),
                "age_min_days": float(pd.to_numeric(sub["benchmark_age_days"], errors="coerce").min()),
                "age_max_days": float(pd.to_numeric(sub["benchmark_age_days"], errors="coerce").max()),
            }
        )

    return {
        "overview": pd.DataFrame(summary_rows),
        "all_mapped_images": df_all,
        "oct_agepred_eligible": eligible,
        "paper1_controls_day0_day90": controls,
        "paper1_controls_hls_day0_day90": allgroups,
    }


def build_column_dictionary() -> pd.DataFrame:
    rows = [
        ("image_path_relative", "Path to the source image relative to the repository root / local dataset root."),
        ("rat_id", "Rat identifier used for rat-level splitting and aggregation."),
        ("sample_id", "Sample identifier from the metadata mapping file."),
        ("eye", "Eye side: OD or OS."),
        ("cohort", "Numeric cohort label (1, 2, or 3)."),
        ("cohort_label", "Human-readable cohort label: Young Male, Young Female, or Older Male."),
        ("group", "Experimental group derived from the path metadata (Controls, HLS (U), or Baseline)."),
        ("day", "Experimental day label used in the study (for example 0 or 90)."),
        ("chronological_age_days", "Final chronological age in days used as the regression target."),
        ("benchmark_age_days", "Chronological age in days after combining base age with the benchmark day label used in the paper protocol."),
        ("base_age_days", "Baseline age in days before adding the day offset."),
        ("day_offset_days", "Experimental day offset added to the baseline age."),
        ("sex", "Reported sex in the metadata."),
        ("oct_image_type", "Mapped image type (for example BScanThumb or REGAVG)."),
        ("eye_label", "Material type / eye description from the source metadata."),
        ("session_label", "High-level session label derived from the path (baseline, end_hls, recovery)."),
        ("folder_day", "Raw day value parsed from the folder name."),
        ("bioptigen_subfolder", "Raw Bioptigen subfolder label from the path."),
        ("class_label", "Path-derived class label such as day_0 or day_90."),
        ("control_cv_fold", "Rat-level fold assignment for Controls in the 3-fold primary benchmark; blank for non-Control rows."),
    ]
    return pd.DataFrame(rows, columns=["column_name", "description"])


def load_table(path: Path) -> pd.DataFrame:
    return pd.read_csv(path)


def clean_best_worst_tables() -> Dict[str, pd.DataFrame]:
    manifest = pd.read_csv(BEST_WORST_MANIFEST).copy()
    images = pd.read_csv(BEST_WORST_IMAGES).copy()
    review = pd.read_csv(BEST_WORST_REVIEW).copy()

    if "image_path" in images.columns:
        images["image_path_relative"] = images["image_path"].map(relpath_or_keep)
    if "sample_dir" in review.columns:
        review["sample_dir_relative"] = review["sample_dir"].map(relpath_or_keep)

    best = manifest[manifest["bucket"] == "best"].copy()
    worst = manifest[manifest["bucket"] == "worst"].copy()

    return {
        "best_samples": best,
        "worst_samples": worst,
        "selected_sample_manifest": manifest,
        "selected_images": images,
        "review_index": review,
    }


def polish_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row >= 2:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        ws.sheet_view.showGridLines = True
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            values = ["" if c.value is None else str(c.value) for c in column_cells[:200]]
            max_len = max((len(v) for v in values), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)
    wb.save(path)


def write_workbook(path: Path, sheets: List[Tuple[str, pd.DataFrame]]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name[:31], index=False)
    polish_workbook(path)


def copy_csv(df: pd.DataFrame, name: str) -> None:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(CSV_DIR / name, index=False)


def build_readme(files: List[Dict[str, str]]) -> str:
    lines = [
        "# OSD-679 Age-Prediction Reproducibility Bundle",
        "",
        "This folder contains a polished, GitHub-friendly supplementary data package for the Brown Norway rat retinal age-prediction experiments.",
        "",
        "It is designed to support manuscript reproduction without redistributing the raw OSD-679 image payload.",
        "",
        "## Contents",
        "",
    ]
    for item in files:
        lines.append(f"- `{item['path']}`: {item['description']}")
    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- Raw OCT images are not redistributed here. `image_path_relative` values are pointers into a local OSD-679-style directory layout.",
            "- The primary benchmark subset corresponds to Cohorts 1-3, Controls only, image types `BScanThumb` + `REGAVG`, and study days 0 and 90.",
            "- The broader `Controls + HLS (U)` subset is included because it underlies the control-vs-stress evaluation universe.",
            "- `chronological_age_days` preserves the raw metadata-derived age, whereas `benchmark_age_days` reflects the age implied by the benchmark day label used in the paper protocol.",
            "- Rat-level cross-validation folds are provided for the primary control benchmark.",
            "- The scratch/random ViT baseline is retained in the supplementary result tables as a negative-control architecture check only.",
            "",
            "## Data access",
            "",
            "OSD-679 data access should be requested via NASA GeneLab / the Open Science Data Repository. This repository only provides the derived mapping tables, split definitions, and result summaries used in the paper.",
            "",
            "## Regeneration",
            "",
            "This bundle is generated from the local metadata/results cache with:",
            "",
            "`python3 scripts/paper/build_reproducibility_bundle.py`",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CSV_DIR.mkdir(parents=True, exist_ok=True)

    df_all = clean_image_age_mapping()
    folds = load_table(S9)
    subsets = build_primary_subsets(df_all, folds)
    dictionary = build_column_dictionary()

    wb1_sheets = [
        ("overview", subsets["overview"]),
        ("all_mapped_images", subsets["all_mapped_images"]),
        ("oct_agepred_eligible", subsets["oct_agepred_eligible"]),
        ("paper1_controls_d0d90", subsets["paper1_controls_day0_day90"]),
        ("paper1_allgroups_d0d90", subsets["paper1_controls_hls_day0_day90"]),
        ("column_dictionary", dictionary),
    ]
    write_workbook(WORKBOOK_1, wb1_sheets)

    wb2_sheets = [
        ("cohort_counts_S1", load_table(S1)),
        ("control_cv_S2", load_table(S2)),
        ("control_by_cohort_day_S3", load_table(S3)),
        ("backbone_ablation_main", load_table(TABLE3_MAIN)),
        ("backbone_ablation_full_S4", load_table(S4)),
        ("lora_ablation_S5", load_table(S5)),
        ("distillation_S6", load_table(S6)),
        ("inter_eye_S7", load_table(S7)),
        ("optuna_best_S8", load_table(S8)),
        ("cv_split_rats_S9", load_table(S9)),
    ]
    write_workbook(WORKBOOK_2, wb2_sheets)

    best_worst = clean_best_worst_tables()
    wb3_sheets = [
        ("best_samples", best_worst["best_samples"]),
        ("worst_samples", best_worst["worst_samples"]),
        ("selected_sample_manifest", best_worst["selected_sample_manifest"]),
        ("selected_images", best_worst["selected_images"]),
        ("review_index", best_worst["review_index"]),
    ]
    write_workbook(WORKBOOK_3, wb3_sheets)

    copy_csv(subsets["all_mapped_images"], "osd679_c123_all_mapped_images_minimal.csv")
    copy_csv(subsets["paper1_controls_day0_day90"], "osd679_paper1_controls_day0_day90_manifest.csv")
    copy_csv(subsets["paper1_controls_hls_day0_day90"], "osd679_paper1_controls_hls_day0_day90_manifest.csv")
    copy_csv(load_table(S9), "osd679_paper1_control_cv_fold_definitions.csv")
    copy_csv(load_table(S3), "osd679_paper1_control_performance_by_cohort_day.csv")
    copy_csv(load_table(TABLE3_MAIN), "osd679_paper1_backbone_ablation_mainpaper.csv")
    copy_csv(best_worst["selected_sample_manifest"], "osd679_paper1_best_worst_control_examples.csv")

    file_entries = [
        {
            "path": str(path.relative_to(ROOT)),
            "description": desc,
        }
        for path, desc in [
            (WORKBOOK_1, "Excel workbook with the full image-to-age mapping, OCT-eligible subsets, and the primary day 0/90 manifests."),
            (WORKBOOK_2, "Excel workbook with cohort counts, benchmark results, split definitions, and supplementary tables."),
            (WORKBOOK_3, "Excel workbook with best/worst qualitative example metadata and image-level review indices."),
            (CSV_DIR / "osd679_c123_all_mapped_images_minimal.csv", "CSV companion for the cleaned Cohort 1-3 image-to-age mapping."),
            (CSV_DIR / "osd679_paper1_controls_day0_day90_manifest.csv", "CSV companion for the primary control day 0/90 benchmark manifest."),
            (CSV_DIR / "osd679_paper1_controls_hls_day0_day90_manifest.csv", "CSV companion for the broader Controls + HLS (U) day 0/90 evaluation universe."),
            (CSV_DIR / "osd679_paper1_control_cv_fold_definitions.csv", "Rat-level 3-fold cross-validation definitions for the primary control benchmark."),
            (CSV_DIR / "osd679_paper1_control_performance_by_cohort_day.csv", "Per-cohort, per-day control performance table used in the manuscript."),
            (CSV_DIR / "osd679_paper1_backbone_ablation_mainpaper.csv", "Main-text backbone ablation table (RETFound + LoRA vs Xception + GAP)."),
            (CSV_DIR / "osd679_paper1_best_worst_control_examples.csv", "Best/worst qualitative sample manifest for the Xception control review set."),
        ]
    ]

    README_MD.write_text(build_readme(file_entries), encoding="utf-8")

    manifest = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_files": {
            "image_age_mapping": str(IMAGE_AGE_MAP.relative_to(ROOT)),
            "supplementary_s1": str(S1.relative_to(ROOT)),
            "supplementary_s9": str(S9.relative_to(ROOT)),
            "qualitative_manifest": str(BEST_WORST_MANIFEST.relative_to(ROOT)),
        },
        "primary_counts": subsets["overview"].to_dict(orient="records"),
        "files": file_entries,
    }
    MANIFEST_JSON.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps({"out_dir": str(OUT_DIR), "files": file_entries}, indent=2))


if __name__ == "__main__":
    main()
